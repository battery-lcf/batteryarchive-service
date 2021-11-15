import os
import glob
import pandas as pd
from google.cloud import storage
from pathlib import Path
import shutil
import openpyxl
import zipfile
#import pyexcel as p

     
def upload_blob(bucket_name, directory_path, destination_blob_name, singles):
    """Uploads a folder to the bucket."""
    # The ID of your GCS bucket
    # bucket_name = "your-bucket-name"
    # The path to your file to upload
    # source_file_name = "local/path/to/file"
    # The ID of your GCS object
    # destination_blob_name = "storage-object-name"

    storage_client = storage.Client()
#    storage_client = storage.Client.create_anonymous_client()
    print(storage_client.project)
    bucket = storage_client.bucket(bucket_name)
    print(bucket)
    rel_paths = []
    if(singles):
        rel_paths = glob.glob(os.path.join(directory_path, "**"), recursive=True)
        rel_paths = [path for path in rel_paths if ("Upload" in path or "cell_list" in path)]
    else:
        rel_paths = glob.glob(os.path.join(directory_path, "**"), recursive=True)
        rel_paths = [rel for rel in rel_paths if ("Upload" in rel or "cell_list" in rel)]
    print("Rel_paths: ", rel_paths)
    print("\n")
    for local_file in rel_paths:
        p = Path(local_file)
        partss = p.parts
        print(partss)
        if(partss[-1] == "Upload"):
            continue
        for index, part in enumerate(partss):
            if (part == os.path.split(directory_path)[1]):
                needed = partss[index+1:]
                if(singles):
                    end = [p for p in needed if "Upload" not in p]
                    print(end)
                elif(not singles):
                    end = [p.split()[1] if ("Upload" in p) else p for p in needed]
                    print(end)
        remote_path = f'{destination_blob_name}/{"/".join(end)}'
        print(remote_path)
        if os.path.isfile(local_file):
            blob = bucket.blob(remote_path)
            blob.upload_from_filename(local_file)
    print(
        "Folder {} uploaded to {}.".format(
            directory_path, bucket_name
        )
    )


#get the abosulte path to the celllist
#make the cellist pandas data frame
#unzip each file folder matching in the cell list
#for each excel file in zip folder chech columns

#TODO:stop program if column is missing do not want to upload invalid data
def checkupload(abspath_celllist):
    all_good = True
    single_files = False
    necessary_columns = ["Cycle_Index","Test_Time(s)", "Current(A)", "Voltage(V)", "Date_Time"]
    necessary_columns_csvs = ["Time",   "Voltage",   "Current",  "Temperature", "Date_Time"]
    if os.path.exists(abspath_celllist):
        directory = os.path.split(abspath_celllist)[0]
        print(directory)
        
        #immediately unpack and delete any zips
        any_zips = glob.glob(os.path.join(directory,"*.zip"))
        if(len(any_zips) != 0):
            for zip in any_zips:
                with zipfile.ZipFile(zip) as file:
                    new_dir = os.path.join(directory, os.path.splitext(zip)[0])
                    os.mkdir(new_dir)
                    file.extractall(new_dir)
                os.remove(zip)
        
        df_cellist = pd.read_excel(abspath_celllist)
        print(df_cellist.head)
        
        #these folders could also be individual excel files in rare cases
        folders = glob.glob(os.path.join(directory,"*"))
        folders.remove(abspath_celllist)
        
        print("folders:")
        print(folders)
    
        folder_nums = [os.path.split(folder)[1] for folder in folders]
        print(folder_nums)
        #if there is an extnesion remove it
        
        folder_nums = [os.path.splitext(name)[0] for name in folder_nums]
        print("FOLDERS", folder_nums)

        for index, file_id in enumerate(df_cellist["file_id"]):
            print("FILE_ID", file_id)
            if(file_id in folder_nums):
                #test to see whether folders or individual files
                print("DIRECTORY", directory)
                extension = ""
                if "file_type" in df_cellist:
                    if(df_cellist["file_type"][index] == "xls" or df_cellist["file_type"][index] == "xlsx"):
                        extension_wildcard = "*.xls*"
                        extension = ".xlsx"
                    elif(df_cellist["file_type"][index] == "csv"):
                        extension_wildcard = "*.csv"
                        extension = ".csv"
                else:
                    extension_wildcard = "*.xls*"
                    extension = ".xlsx"

                files = glob.glob(os.path.join(directory,extension_wildcard))
                files.remove(abspath_celllist)
                
                if(len(files) == len(folders)): #means all excel files
                    print("All files are individual files")
                    data_paths = files
                    single_files = True
                    os.mkdir("Upload") #make dir to store all processed files
                    for ind, data in enumerate(data_paths):
                        new_path = os.path.join(directory, "Upload", os.path.split(data)[1])
                        if(os.path.splitext(os.path.split(data)[1])[1] == ".xls"):
#                            p.save_book_as(file_name=data,
#                   dest_file_name= os.path.splitext(new_path)[0]+".xlsx")
                            dest_file_name= os.path.splitext(new_path)[0]+".xlsx"
                            df_convert = pd.read_excel(data, sheet_name=None)
                            dict_keys = list(df_convert.keys())
                            print("Save Sheets: ", dict_keys)
                            #manually create first workbook
                            writer_temp_save1 = pd.ExcelWriter(dest_file_name,engine='openpyxl')
                            sheet1 = dict_keys[0]
                            df_convert[sheet1].to_excel(writer_temp_save1, sheet1, index=False)
                            writer_temp_save1.save()
                            print(f"copying worksheet {sheet1} now")
                            writer_temp_save = pd.ExcelWriter(dest_file_name,engine='openpyxl',mode="a")
                            dict_keys.remove(sheet1)
                            for key in dict_keys:
                                df_convert[key].to_excel(writer_temp_save, key, index=False)
                                print(f"copying worksheet {key} now")
                            
                            writer_temp_save.save()
#                            writer_temp_save.save()
#                            with pd.ExcelWriter(dest_file_name, engine='openpyxl', mode='a') as writer1:
#                                for key in df_convert.keys():
#                                    df_convert[key].to_excel(writer1, sheet_name = key, index = False)
                            print("converted xls to xlsx for uploaded files")
                        else:
                            shutil.copy(data, new_path)
                        new_path = os.path.splitext(new_path)[0]+extension
                        data_paths[ind] = new_path #replace the path with copy
                        
                else:
                    print("All excel files are in folders")
                    data_paths = glob.glob(os.path.join(directory,file_id,extension_wildcard))
                    print(data_paths)
                    os.mkdir("Upload "+str(file_id))
                    for ind, data in enumerate(data_paths):
                        new_path = os.path.join(directory, "Upload "+str(file_id), os.path.split(data)[1])
                        if(os.path.splitext(os.path.split(data)[1])[1] == ".xls"):
#                            p.save_book_as(file_name=data,
#                   dest_file_name= os.path.splitext(new_path)[0]+".xlsx")
                            dest_file_name= os.path.splitext(new_path)[0]+".xlsx"
                            df_convert = pd.read_excel(data, sheet_name=None)
                            dict_keys = df_convert.keys()
                            
                            writer_temp_save1 = pd.ExcelWriter(dest_file_name,engine='openpyxl')
                            sheet1 = dict_keys[0]
                            df_convert[sheet1].to_excel(writer_temp_save1, sheet1, index=False)
                            writer_temp_save1.save()
                            print(f"copying worksheet {sheet1} now")
                            writer_temp_save = pd.ExcelWriter(dest_file_name,engine='openpyxl',mode="a")
                            dict_keys.remove(sheet1)
                            for key in dict_keys:
                                df_convert[key].to_excel(writer_temp_save, key, index=False)
                                print(f"copying worksheet {key} now")
                            writer_temp_save.save()
                            
                        else:
                            shutil.copy(data, new_path)
                        new_path = os.path.splitext(new_path)[0]+extension
                        data_paths[ind] = new_path #replace the path with copy
                        
                for file in data_paths:
                    print("File:" + file)
                    print("Checking file:" + os.path.split(file)[1])
                    #check columns and spreadsheet name
                    if(extension == ".xlsx"):
                        sheets_dict = pd.read_excel(file, sheet_name=None) #creates dict of
                        sheetnames = sheets_dict.keys()
                        print(sheetnames)
                        ind = False
                        writer = pd.ExcelWriter(file,engine='openpyxl', mode="a", if_sheet_exists="replace")
                        for k in sheetnames: #check valid cols and Date_time
                            if "hannel" in k:
                                    
                                ind = True
                                df_hannel = sheets_dict[k]
            
                                if('Date_Time' not in df_hannel.columns):
                                        
                                        df_hannel['Date_Time'] = pd.to_datetime(df_hannel['Test_Time(s)'], unit="s")
                                        
                                        
                                        df_hannel.to_excel(writer, sheet_name=k, index=False)
                                        print("Date_Time ADDED")
                                if(set(necessary_columns).issubset(df_hannel.columns)):
                                        print("Channel worksheet has all valid columns")
                                else:
                                        print("MISSING: Channel worksheet in file " + file + " does not have all valid columns")
                                        all_good = False
                        writer.save()
                    elif extension == ".csv":
                        df_hannel = pd.read_csv(file)
                        if('Date_Time' not in df_hannel.columns):
                                df_hannel['Date_Time'] = pd.to_datetime(df_hannel['Test_Time(s)'], unit="s")
                                df_hannel.to_csv(file)
                                print("Date_Time ADDED")
                        if(set(necessary_columns_csvs).issubset(df_hannel.columns)):
                                print("csv has all valid columns")
                        else:
                                print("MISSING: csv " + file + " does not have all valid columns")
                                all_good = False
                        

                    if(ind == False):
                        print("There is no worksheet named channel in " + file + " please change the worksheet name. Please rename correct worksheet to channel and data time will be added if necessary")
                        all_good = False
                if(single_files):
                    print("done processing single excel files")
                    break
                


            else:
                print("There is a file_id in cell_list that is not present as a folder or file in the directory")
                all_good = False
    else:
        print("Invalid absolute path. Are you sure this is where the cell_list file is located?")
        all_good = False
    return all_good, single_files



if __name__ == "__main__":
    #should be running from the scripts folder
    print(os.getcwd())
    while True:
        folder_path = input("What is the full path of folder you want to upload data from? (Drag and drop folder into terminal for full path) ")
        if(os.path.exists(folder_path)):
            os.chdir(folder_path)
            break
        else:
            print("That is not the right path. Please re-enter the correct path")
    
    
    pwd = os.getcwd()
    print(pwd)
    if(os.path.exists(pwd)):
        contents = os.listdir(pwd)
        print(contents)
        cell_list_there = False
        for item in contents:
            if "cell_list.xls" in item:
                abspath_list = glob.glob(os.path.join(pwd,"cell_list.xls*"))
                print(abspath_list)
                print("FOUND cell_list")
                cell_list_there = True
        if(cell_list_there == False):
                print("MISSING: cell_list")
                quit()
            
    #verify folder to upload
    status, singles = checkupload(abspath_list[0])
#    status = True
#    singles = False
    
    if(status == False):
        print("Please fix according to the error messages above and then rerun to upload")
        quit()
#ba_samplefile
    bucket_name = "ba_datasets"
    local_path = pwd
    destination_blob_name = os.path.split(pwd)[1]
    #upload folder
    print(destination_blob_name)
    upload_blob(bucket_name, local_path, destination_blob_name, singles)
    
    contents = os.listdir(pwd)
    temp_upload_folders = [fold for fold in contents if "Upload" in fold]
    [shutil.rmtree(os.path.abspath(f)) for f in temp_upload_folders]
    print("Deleted Temporary folders")
